from bs4 import BeautifulSoup
import xlwt
import os
import re
import copy
import openpyxl

from transfer import transfer

# input_object = {'"Analytic Methods in Accident Research"': 5, '"COMPUTER-AIDED CIVIL AND INFRASTRUCTURE ENGINEERING"': 5, '"IEEE Vehicular Technology Magazine"': 5, '"TRANSPORT REVIEWS"': 5, '"IEEE TRANSACTIONS ON INTELLIGENT TRANSPORTATION SYSTEMS"': 5, '"TRANSPORTATION RESEARCH PART C-EMERGING TECHNOLOGIES"': 5, '"IEEE Transactions on Transportation Electrification"': 5, '"IEEE TRANSACTIONS ON VEHICULAR TECHNOLOGY"': 5, '"TRANSPORTATION RESEARCH PART B-METHODOLOGICAL"': 5, '"Vehicular Communications"': 5, '"TRANSPORTATION RESEARCH PART E-LOGISTICS AND TRANSPORTATION REVIEW"': 5, '"TRANSPORTATION RESEARCH PART D-TRANSPORT AND ENVIRONMENT"': 5, '"TRANSPORTATION"': 5, '"TRANSPORTATION RESEARCH PART A-POLICY AND PRACTICE"': 5, '"Journal of Transport Geography"': 5, '"ACCIDENT ANALYSIS AND PREVENTION"': 5}
input_object = transfer()
print(len(input_object))
title_set = set()
jou_set = set()

def write_list_info():
    workbook = openpyxl.Workbook()
    # workbook = xlwt.Workbook(encoding='utf-8')
    # worksheet = workbook.add_sheet('info')
    worksheet = workbook.active
    # worksheet.cell(1,0+1, '期刊/会议')
    worksheet.cell(1,1, '期刊/会议名字')
    worksheet.cell(1,2, '年份')
    worksheet.cell(1,3, '作者')
    worksheet.cell(1,4, '论文名称')

    ConOrJou_list = [x.replace('"','').lower() for x in list(input_object.keys())]

    index = 1
    sheet_index = 1
    file_list = os.listdir('sci_download')
    for file_name in file_list:
        with open(f'sci_download/{file_name}','r',encoding='utf-8') as f:
            html_doc = f.read()

        soup = BeautifulSoup(html_doc, 'html.parser')
        for table in soup.find_all('table'):
            if table.tr.td.text != 'PT ':
                continue
            for tr in table.find_all('tr'):
                if tr.td.text == 'AU ':
                    authors = [x.strip() for x in tr.find_all('td')[1].get_text().split('\n')]
                elif tr.td.text == 'TI ':
                    title = ''.join([x.strip() for x in tr.find_all('td')[1].text.split('\n')])
                elif tr.td.text == 'SO ':
                    ConOrJou = ' '.join([x.strip() for x in tr.find_all('td')[1].text.lower().split('\n')])
                elif tr.td.text == 'PY ':
                    year = tr.find_all('td')[1].text
            # jou_set.add(ConOrJou)
            if ConOrJou.lower() not in ConOrJou_list: 
                continue
            if title not in title_set:
                title_set.add(title)
            else:
                continue
            jou_set.add(ConOrJou)

            if index >= 100000:
                index = 1
                workbook.create_sheet(f'Sheet{sheet_index}')
                worksheet = workbook.get_sheet_by_name(f'Sheet{sheet_index}')
                # worksheet.cell(1,0+1, '期刊/会议')
                worksheet.cell(1,1, '期刊/会议名字')
                worksheet.cell(1,2, '年份')
                worksheet.cell(1,3, '作者')
                worksheet.cell(1,4, '论文名称')
                sheet_index += 1

            worksheet.cell(index+1,2-1,ConOrJou)
            worksheet.cell(index+1,3-1,year)
            worksheet.cell(index+1,4-1,','.join(authors))
            worksheet.cell(index+1,5-1,title)
            index += 1

    workbook.save('sci_list_info.xlsx')

def write_sort_info():
    author_set = set()
    author_list = [x.replace('"','').lower() for x in list(input_object.keys())]
    # s = ' '.join(ConOrJou_list)
    # print(s)
    file_dir = 'sci_download'
    file_list = os.listdir(file_dir)
    # init xls

    result_as_authors = dict()

    for file_name in file_list:
        with open(f'{file_dir}/{file_name}','r',encoding='utf-8') as f:
            html_doc = f.read()
        print(file_name)
        soup = BeautifulSoup(html_doc, 'html.parser')

        authors = []
        title = ''
        ConOrJou = ''
        year = ''

        for table in soup.find_all('table'):
            if table.tr.td.text != 'PT ':
                continue
            for tr in table.find_all('tr'):
                if tr.td.text == 'AU ':
                    authors = [x.strip() for x in tr.find_all('td')[1].get_text().split('\n')]
                elif tr.td.text == 'TI ':
                    title = ''.join([x.strip() for x in tr.find_all('td')[1].text.split('\n')])
                elif tr.td.text == 'SO ':
                    ConOrJou = ' '.join([x.strip() for x in tr.find_all('td')[1].text.lower().split('\n')])
                elif tr.td.text == 'PY ':
                    year = tr.find_all('td')[1].text
            
            for index,author in enumerate(authors):
                if author.lower() not in author_list: # filter out the author in author list
                    continue
                author_set.add(author)
                if author not in result_as_authors:
                    result_as_authors[author] = dict()
                    if index == 0:
                        result_as_authors[author]['1st_count'] = 1
                        result_as_authors[author]['not_1st_count'] = 0
                        result_as_authors[author]['flag'] = ['True']
                    else:
                        result_as_authors[author]['1st_count'] = 0
                        result_as_authors[author]['not_1st_count'] = 1
                        result_as_authors[author]['flag'] = ['False']
                    result_as_authors[author]['title'] = [title]
                    # result_as_authors[author]['ConOrJou'] = [SN_LIST[N_LIST.index(ConOrJou)]]
                    result_as_authors[author]['ConOrJou'] = [ConOrJou]
                    result_as_authors[author]['year'] = [year]
                    result_as_authors[author]['n_list'] = [copy.deepcopy(authors)]
                    result_as_authors[author]['co-worker'] = copy.deepcopy(authors)
                    result_as_authors[author]['co-worker'].remove(author)
                else:
                    if title in result_as_authors[author]['title']:
                        continue
                    if index == 0:
                        result_as_authors[author]['1st_count'] += 1
                        result_as_authors[author]['flag'].append('True')
                    else:
                        result_as_authors[author]['not_1st_count'] += 1
                        result_as_authors[author]['flag'].append('False')
                    result_as_authors[author]['title'].append(title)
                    # result_as_authors[author]['ConOrJou'].append(SN_LIST[N_LIST.index(ConOrJou)])
                    result_as_authors[author]['ConOrJou'].append(ConOrJou)
                    result_as_authors[author]['year'].append(year)
                    result_as_authors[author]['n_list'].append(copy.deepcopy(authors))
                    result_as_authors[author]['co-worker'] += copy.deepcopy(authors)
                    result_as_authors[author]['co-worker'].remove(author)

            authors = []
            title = ''
            ConOrJou = ''

    print('start2write')
    print(len(author_set),author_set)
    print(result_as_authors.keys())
            # 这里还要插一个对result_as_authors的排序
    result_as_authors_list = sorted(result_as_authors.items(), key = lambda x : x[1]['1st_count'] + x[1]['not_1st_count'], reverse = True)


    workbook = openpyxl.Workbook()
    # workbook = xlwt.Workbook(encoding='utf-8')
    # worksheet = workbook.add_sheet('info')
    worksheet = workbook.active
    worksheet.cell(1,0+1, '作者')
    worksheet.cell(1,1+1, '第一作者paper数量')
    worksheet.cell(1,2+1, '非第一作者paper数量')
    worksheet.cell(1,3+1, 'title')
    worksheet.cell(1,4+1, '会议\期刊名')
    worksheet.cell(1,5+1, '年份')
    worksheet.cell(1,6+1, '是否一作')
    worksheet.cell(1,7+1, '本文所有作者')
    worksheet.cell(1,9+1, 'co-worker')
    worksheet.cell(1,10+1, '合作次数')

    # sheet_index = 1
    book_index = 1
    index = 2

    for content in result_as_authors_list:
        # get rid of those researcher with only one paper
        # if content[1]['1st_count'] + content[1]['not_1st_count'] <= 5:
        #     continue
        if index > 1000000:
            workbook.save(f'sci_sort_info{book_index}.xlsx')
            workbook = openpyxl.Workbook()
            # workbook = xlwt.Workbook(encoding='utf-8')
            # worksheet = workbook.add_sheet('info')
            worksheet = workbook.active
            worksheet.cell(1,0+1, '作者')
            worksheet.cell(1,1+1, '第一作者paper数量')
            worksheet.cell(1,2+1, '非第一作者paper数量')
            worksheet.cell(1,3+1, 'title')
            worksheet.cell(1,4+1, '会议\期刊名')
            worksheet.cell(1,5+1, '年份')
            worksheet.cell(1,6+1, '是否一作')
            worksheet.cell(1,7+1, '本文所有作者')
            worksheet.cell(1,9+1, 'co-worker')
            worksheet.cell(1,10+1, '合作次数')
            index = 2
            book_index += 1
            # workbook.create_sheet(f'Sheet{sheet_index}')
            # worksheet = workbook.get_sheet_by_name(f'Sheet{sheet_index}')
            # worksheet.cell(1,0+1, '作者')
            # worksheet.cell(1,1+1, '第一作者paper数量')
            # worksheet.cell(1,2+1, '非第一作者paper数量')
            # worksheet.cell(1,3+1, 'title')
            # worksheet.cell(1,4+1, '会议\期刊名')
            # worksheet.cell(1,5+1, '年份')
            # worksheet.cell(1,6+1, '是否一作')
            # worksheet.cell(1,7+1, '本文所有作者')
            # worksheet.cell(1,9+1, 'co-worker')
            # worksheet.cell(1,10+1, '合作次数')
            # index = 2 
            # sheet_index += 1
        if content[0] == '[Anonymous]':
            continue
        worksheet.cell(index, 0+1,  content[0])
        worksheet.cell(index, 1+1,  content[1]['1st_count'])
        worksheet.cell(index, 2+1,  content[1]['not_1st_count'])
        temp = index
        flag = index
        for i, article in enumerate(content[1]['title']):
            worksheet.cell(temp, 3+1,  article)
            worksheet.cell(temp, 4+1,  content[1]['ConOrJou'][i])
            worksheet.cell(temp, 5+1,  content[1]['year'][i])
            worksheet.cell(temp, 6+1,  content[1]['flag'][i])
            worksheet.cell(temp, 7+1,  '; '.join(content[1]['n_list'][i]))
            temp += 1
        cw_dict = dict()
        for item in content[1]['co-worker']:
            cw_dict[item] = cw_dict.get(item, 0) + 1
        cw_list = sorted(cw_dict.items(), key = lambda x : x[1], reverse = True)
        col = 9 + 1
        num_cw = 0
        for cw in cw_list:
            worksheet.cell(index, col,  cw[0])
            worksheet.cell(index, col+1,  cw[1])
            index += 1
            num_cw += 1
            if num_cw > 500:
                break
            # if index == temp:
            #     col += 2
            #     index = flag
            #     worksheet.cell(1,col, 'co-worker')
            #     worksheet.cell(1,col+1, '合作次数')
        index = max(temp, index)
    if book_index == 1:
        workbook.save('sci_sort_info_author.xlsx')

# write_list_info()
# print(jou_set)
write_sort_info()