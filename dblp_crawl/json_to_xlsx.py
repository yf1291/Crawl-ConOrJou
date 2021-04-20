import xlwt
import json
import re
import openpyxl
import copy

def get_year(data):
    pattern = re.compile(r'20\d\d')
    time_tmp = pattern.findall(data)
    if len(time_tmp) > 0:
        year = time_tmp[0]
    else:
        year = None
    return year 

def write_list_info():
    workbook = openpyxl.Workbook()
    # workbook = xlwt.Workbook(encoding='utf-8')
    # worksheet = workbook.add_sheet('info')
    worksheet = workbook.active
    worksheet.cell(1,0+1, '期刊/会议')
    worksheet.cell(1,1+1, '期刊/会议名字')
    worksheet.cell(1,2+1, '年份')
    worksheet.cell(1,3+1, '作者')
    worksheet.cell(1,4+1, '论文名称')

    with open('info.json','r') as f:
        info_dict = json.load(f)

    index = 1
    sheet_index = 1
    for info in info_dict:
        if index >= 100000:
            index = 1
            workbook.create_sheet(f'Sheet{sheet_index}')
            worksheet = workbook.get_sheet_by_name(f'Sheet{sheet_index}')
            worksheet.cell(1,0+1, '期刊/会议')
            worksheet.cell(1,1+1, '期刊/会议名字')
            worksheet.cell(1,2+1, '年份')
            worksheet.cell(1,3+1, '作者')
            worksheet.cell(1,4+1, '论文名称')
            sheet_index += 1
        year = get_year(info['ConOrJou'])
        if not year:
            continue
        worksheet.cell(index+1,1,info['ConOrJou'])
        worksheet.cell(index+1,2,info['ConOrJouName'])
        worksheet.cell(index+1,3,year)
        worksheet.cell(index+1,4,','.join(info['authors']))
        worksheet.cell(index+1,5,info['title'])
        index += 1

    workbook.save('dblp_list_info.xlsx')

def write_sort_info():
    with open('info.json','r') as f:
        info_dict = json.load(f)

    result_as_authors = dict()
    authors = []
    title = ''
    ConOrJou = ''
    year = ''

    for info in info_dict:
        # 匹配20开头的数字
        year = get_year(info['ConOrJou'])
        if not year:
            continue
        authors = info['authors']
        title = info['title']
        ConOrJou = info['ConOrJouName']

        flag = True
        for author in authors:
            if author not in result_as_authors:
                result_as_authors[author] = dict()
                if flag:
                    result_as_authors[author]['1st_count'] = 1
                    result_as_authors[author]['not_1st_count'] = 0
                    result_as_authors[author]['flag'] = ['True']
                else:
                    result_as_authors[author]['1st_count'] = 0
                    result_as_authors[author]['not_1st_count'] = 1
                    result_as_authors[author]['flag'] = ['False']
                result_as_authors[author]['title'] = [title]
                # result_as_authors[author]['ConOrJou'] = [SN_LIST[N_LIST.index(ConOrJou)]]
                result_as_authors[author]['ConOrJou'] = [ConOrJou]
                result_as_authors[author]['year'] = [year]
                result_as_authors[author]['n_list'] = [copy.deepcopy(authors)]
                result_as_authors[author]['co-worker'] = copy.deepcopy(authors)
                result_as_authors[author]['co-worker'].remove(author)
            else:
                if title in result_as_authors[author]['title']:
                    continue
                if flag:
                    result_as_authors[author]['1st_count'] += 1
                    result_as_authors[author]['flag'].append('True')
                else:
                    result_as_authors[author]['not_1st_count'] += 1
                    result_as_authors[author]['flag'].append('False')
                result_as_authors[author]['title'].append(title)
                # result_as_authors[author]['ConOrJou'].append(SN_LIST[N_LIST.index(ConOrJou)])
                result_as_authors[author]['ConOrJou'].append(ConOrJou)
                result_as_authors[author]['year'].append(year)
                result_as_authors[author]['n_list'].append(copy.deepcopy(authors))
                result_as_authors[author]['co-worker'] += copy.deepcopy(authors)
                result_as_authors[author]['co-worker'].remove(author)
            flag = False

        authors = []
        title = ''
        ConOrJou = ''

            # 这里还要插一个对result_as_authors的排序
    result_as_authors_list = sorted(result_as_authors.items(), key = lambda x : x[1]['1st_count'] + x[1]['not_1st_count'], reverse = True)

    workbook = openpyxl.Workbook()
    # workbook = xlwt.Workbook(encoding='utf-8')
    # worksheet = workbook.add_sheet('info')
    worksheet = workbook.active
    worksheet.cell(1,0+1, '作者')
    worksheet.cell(1,1+1, '第一作者paper数量')
    worksheet.cell(1,2+1, '非第一作者paper数量')
    worksheet.cell(1,3+1, 'title')
    worksheet.cell(1,4+1, '会议\期刊名')
    worksheet.cell(1,5+1, '年份')
    worksheet.cell(1,6+1, '是否一作')
    worksheet.cell(1,7+1, '本文所有作者')
    worksheet.cell(1,9+1, 'co-worker')
    worksheet.cell(1,10+1, '合作次数')

    sheet_index = 1
    index = 2

    for content in result_as_authors_list:
        if index > 1000000:
            workbook.create_sheet(f'Sheet{sheet_index}')
            worksheet = workbook.get_sheet_by_name(f'Sheet{sheet_index}')
            worksheet.cell(1,0+1, '作者')
            worksheet.cell(1,1+1, '第一作者paper数量')
            worksheet.cell(1,2+1, '非第一作者paper数量')
            worksheet.cell(1,3+1, 'title')
            worksheet.cell(1,4+1, '会议\期刊名')
            worksheet.cell(1,5+1, '年份')
            worksheet.cell(1,6+1, '是否一作')
            worksheet.cell(1,7+1, '本文所有作者')
            worksheet.cell(1,9+1, 'co-worker')
            worksheet.cell(1,10+1, '合作次数')
            index = 2 
            sheet_index += 1
        if content[0] == '[Anonymous]':
            continue
        worksheet.cell(index, 0+1,  content[0])
        worksheet.cell(index, 1+1,  content[1]['1st_count'])
        worksheet.cell(index, 2+1,  content[1]['not_1st_count'])
        temp = index
        flag = index
        for i, article in enumerate(content[1]['title']):
            worksheet.cell(temp, 3+1,  article)
            worksheet.cell(temp, 4+1,  content[1]['ConOrJou'][i])
            worksheet.cell(temp, 5+1,  content[1]['year'][i])
            worksheet.cell(temp, 6+1,  content[1]['flag'][i])
            worksheet.cell(temp, 7+1,  '; '.join(content[1]['n_list'][i]))
            temp += 1
        cw_dict = dict()
        for item in content[1]['co-worker']:
            cw_dict[item] = cw_dict.get(item, 0) + 1
        cw_list = sorted(cw_dict.items(), key = lambda x : x[1], reverse = True)
        col = 9 + 1
        for cw in cw_list:
            worksheet.cell(index, col,  cw[0])
            worksheet.cell(index, col+1,  cw[1])
            index += 1
            # if index == temp:
            #     col += 2
            #     index = flag
            #     worksheet.cell(1,col, 'co-worker')
            #     worksheet.cell(1,col+1, '合作次数')
        index = max(temp, index)

    workbook.save('dblp_sort_info.xlsx')

write_list_info()
write_sort_info()